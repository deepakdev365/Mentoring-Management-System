import openpyxl
import os

def clean_duplicates(file_path):
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    headers = [cell.value for cell in sheet[1]]
    reg_idx = -1
    sub_idx = -1
    sem_idx = -1
    for i, h in enumerate(headers):
        h_low = str(h).lower().strip()
        if h_low in ['registrationnumber', 'rollno', 'registrationno']:
            reg_idx = i + 1
        elif h_low in ['subjectcode', 'code']:
            sub_idx = i + 1
        elif h_low in ['semester', 'sem']:
            sem_idx = i + 1
            
    if reg_idx == -1 or sub_idx == -1 or sem_idx == -1:
        print(f"Columns not found in {file_path}. Headers: {headers}")
        return

    seen = set()
    rows_to_delete = []
    
    # Iterate backwards to delete rows without affecting indices
    for row_num in range(2, sheet.max_row + 1):
        reg = str(sheet.cell(row=row_num, column=reg_idx).value).strip()
        sub = str(sheet.cell(row=row_num, column=sub_idx).value).strip()
        sem = str(sheet.cell(row=row_num, column=sem_idx).value).strip()
        key = f"{reg}-{sub}-{sem}"
        
        if key in seen:
            rows_to_delete.append(row_num)
        else:
            seen.add(key)

    if rows_to_delete:
        print(f"Found {len(rows_to_delete)} duplicate rows in {file_path}. Deleting...")
        # Delete rows from bottom to top
        for row_num in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row_num)
        wb.save(file_path)
        print(f"Cleaned {file_path}.")
    else:
        print(f"No duplicates found in {file_path}.")

files = [
    r'd:\projects\Github\Mentoring-Management-System\mock_data\SubjectRegistration_Mock.xlsx',
    r'd:\projects\Github\Mentoring-Management-System\mock_data\marks_upload.xlsx',
    r'd:\projects\Github\Mentoring-Management-System\mock_data\Attendance_Mock_Data.xlsx'
]

for f in files:
    clean_duplicates(f)
