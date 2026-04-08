import openpyxl
import os

def normalize(val):
    return str(val).lower().replace(" ", "").replace("_", "").replace("-", "").strip()

def clean_to_new_file(file_path):
    if not os.path.exists(file_path): return
    print(f"Cleaning {file_path} to V2...")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        
        reg_idx, code_idx, sem_idx = -1, -1, -1
        for i, h in enumerate(headers):
            h_norm = normalize(h)
            if h_norm in ['registrationnumber', 'rollno', 'registrationno', 'regno']: reg_idx = i
            elif h_norm in ['subjectcode', 'code']: code_idx = i
            elif h_norm in ['semester', 'sem']: sem_idx = i
                
        if reg_idx == -1 or code_idx == -1 or sem_idx == -1: return

        seen = set()
        rows_to_keep = [headers]
        for row_num in range(2, sheet.max_row + 1):
            k = f"{str(sheet.cell(row=row_num, column=reg_idx+1).value).strip()}|{str(sheet.cell(row=row_num, column=code_idx+1).value).strip()}|{str(sheet.cell(row=row_num, column=sem_idx+1).value).strip()}"
            if k not in seen:
                seen.add(k)
                rows_to_keep.append([sheet.cell(row=row_num, column=c+1).value for c in range(len(headers))])

        new_path = file_path.replace(".xlsx", "_Clean.xlsx")
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        for r in rows_to_keep: new_sheet.append(r)
        new_wb.save(new_path)
        print(f"Saved clean version to: {new_path}")
    except Exception as e:
        print(f"Failed: {e}")

files = [
    r'd:\projects\Github\Mentoring-Management-System\mock_data\SubjectRegistration_Mock.xlsx',
    r'd:\projects\Github\Mentoring-Management-System\mock_data\marks_upload.xlsx',
    r'd:\projects\Github\Mentoring-Management-System\mock_data\Attendance_Mock_Data.xlsx',
    r'd:\projects\Github\Mentoring-Management-System\mock_data\Payment_Mock_Data.xlsx'
]

for f in files:
    clean_to_new_file(f)
