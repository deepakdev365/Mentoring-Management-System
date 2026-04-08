import openpyxl
import random
from datetime import datetime, timedelta

def generate_payment_data():
    # 1. Get registration numbers from Students_Mock_Data.xlsx
    students_wb = openpyxl.load_workbook(r'Students_Mock_Data.xlsx')
    students_sheet = students_wb.active
    headers = [cell.value for cell in students_sheet[1]]
    
    reg_idx = -1
    name_idx = -1
    email_idx = -1
    for i, h in enumerate(headers):
        h_low = str(h).lower().strip()
        if h_low in ['registrationnumber', 'registrationno', 'rollno']:
            reg_idx = i + 1
        elif h_low in ['fullname', 'studentname', 'name']:
            name_idx = i + 1
        elif h_low in ['email', 'mail']:
            email_idx = i + 1
    
    if reg_idx == -1: 
        print("Required columns not found in students sheet.")
        return

    student_records = []
    for row in range(2, students_sheet.max_row + 1):
        reg = students_sheet.cell(row=row, column=reg_idx).value
        name = students_sheet.cell(row=row, column=name_idx).value if name_idx != -1 else "Student"
        email = students_sheet.cell(row=row, column=email_idx).value if email_idx != -1 else "email@example.com"
        if reg: student_records.append({'reg': str(reg), 'name': str(name), 'email': str(email)})

    # 2. Generate Payment_Mock_Data.xlsx
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Payments"
    
    # Headers matching UploadPaymentsComponent.ts requirements/aliases
    headers = ["rollNo", "studentName", "email", "amount", "semester", "paymentDate", "status", "mode"]
    sheet.append(headers)

    modes = ["UPI", "CARD", "NETBANKING", "CASH"]
    statuses = ["SUCCESS", "SUCCESS", "SUCCESS", "FAILED", "PENDING"] # Weighted for success

    for record in student_records:
        # Generate 1-2 payments per student
        for _ in range(random.randint(1, 2)):
            amount = random.choice([15000, 25000, 45000, 50000, 75000])
            semester = "3rd"
            # Random date in the last 30 days
            date = datetime.now() - timedelta(days=random.randint(0, 30))
            status = random.choice(statuses)
            mode = random.choice(modes)
            
            sheet.append([
                record['reg'],
                record['name'],
                record['email'],
                amount,
                semester,
                date.strftime("%Y-%m-%d %H:%M:%S"),
                status,
                mode
            ])

    wb.save('Payment_Mock_Data.xlsx')
    print(f"Generated Payment_Mock_Data.xlsx with {sheet.max_row - 1} records.")

if __name__ == "__main__":
    generate_payment_data()
