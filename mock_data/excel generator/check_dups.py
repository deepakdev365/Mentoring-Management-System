import openpyxl
import os

wb = openpyxl.load_workbook('SubjectRegistration_Mock.xlsx')
s = wb.active
h = [cell.value for cell in s[1]]
cols = {str(v).lower().strip(): i for i,v in enumerate(h)}

reg_idx = cols.get('registrationno', cols.get('rollno', cols.get('registrationnumber')))
code_idx = cols.get('subjectcode', cols.get('code'))
sem_idx = cols.get('semester', cols.get('sem'))

seen = set()
dups = []
for r in range(2, s.max_row+1):
    reg = str(s.cell(r, reg_idx+1).value).strip()
    code = str(s.cell(r, code_idx+1).value).strip()
    sem = str(s.cell(r, sem_idx+1).value).strip()
    key = f"{reg}|{code}|{sem}"
    if key in seen:
        dups.append(r)
    else:
        seen.add(key)

print(f"Total Rows: {s.max_row}")
print(f"Duplicates: {len(dups)}")
if dups:
    print(f"First Dup Row: {dups[0]}")
    # Display the dup info
    row = dups[0]
    print(f"Data: {s.cell(row, reg_idx+1).value}, {s.cell(row, code_idx+1).value}, {s.cell(row, sem_idx+1).value}")
