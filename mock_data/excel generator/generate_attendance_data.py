import openpyxl
import random

def generate_attendance():
    # 1. Get registration numbers from Students_Mock_Data.xlsx
    students_wb = openpyxl.load_workbook(r'Students_Mock_Data.xlsx')
    students_sheet = students_wb.active
    headers = [cell.value for cell in students_sheet[1]]
    reg_idx = -1
    for i, h in enumerate(headers):
        if h and str(h).lower().strip() in ['registrationnumber', 'registrationno', 'rollno']:
            reg_idx = i + 1
            break
    
    if reg_idx == -1: 
        print("Registration number column not found in students sheet.")
        return

    reg_numbers = []
    for row in range(2, students_sheet.max_row + 1):
        val = students_sheet.cell(row=row, column=reg_idx).value
        if val: reg_numbers.append(str(val))

    # 2. Get subject codes from SubjectRegistration_Mock.xlsx
    sub_wb = openpyxl.load_workbook(r'SubjectRegistration_Mock.xlsx')
    sub_sheet = sub_wb.active
    sub_headers = [cell.value for cell in sub_sheet[1]]
    sub_idx = -1
    for i, h in enumerate(sub_headers):
        if h and str(h).lower().strip() in ['subjectcode', 'code']:
            sub_idx = i + 1
            break
    
    if sub_idx == -1:
        # Fallback to some common codes
        subject_codes = ["CS101", "MA102", "PH103", "EC201"]
    else:
        subject_codes = list(set([str(sub_sheet.cell(row=row, column=sub_idx).value) for row in range(2, min(50, sub_sheet.max_row + 1)) if sub_sheet.cell(row=row, column=sub_idx).value]))

    # 3. Generate Attendance_Mock_Data.xlsx
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Attendance"
    sheet.append(["registrationNumber", "subjectCode", "percentage", "semester"])

    # For each student, generate attendance for 2-3 subjects
    for reg in reg_numbers:
        # Pick 2-3 random subjects
        chosen_subjects = random.sample(subject_codes, min(len(subject_codes), 3))
        for sub in chosen_subjects:
            percentage = round(random.uniform(75.5, 98.9), 1)
            sheet.append([reg, sub, percentage, "3rd"])

    wb.save('Attendance_Mock_Data.xlsx')
    print(f"Generated Attendance_Mock_Data.xlsx with {sheet.max_row - 1} records.")

if __name__ == "__main__":
    generate_attendance()
