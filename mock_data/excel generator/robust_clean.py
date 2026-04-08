import openpyxl
import os
import shutil

def robust_clean(file_path):
    if not os.path.exists(file_path): return
    
    print(f"Cleaning {file_path}...")
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        
        reg_idx = -1
        sub_idx = -1
        sem_idx = -1
        for i, h in enumerate(headers):
            h_low = str(h).lower().strip()
            if h_low in ['registrationnumber', 'rollno', 'registrationno']: reg_idx = i + 1
            elif h_low in ['subjectcode', 'code']: sub_idx = i + 1
            elif h_low in ['semester', 'sem']: sem_idx = i + 1
                
        if reg_idx == -1 or sub_idx == -1 or sem_idx == -1: 
            print(f"Skipping {file_path} - missing columns.")
            return

        seen = set()
        rows_to_keep = [headers]
        for row_num in range(2, sheet.max_row + 1):
            reg = str(sheet.cell(row=row_num, column=reg_idx).value).strip()
            sub = str(sheet.cell(row=row_num, column=sub_idx).value).strip()
            sem = str(sheet.cell(row=row_num, column=sem_idx).value).strip()
            key = f"{reg}|{sub}|{sem}"
            if key not in seen:
                seen.add(key)
                rows_to_keep.append([sheet.cell(row=row_num, column=c+1).value for c in range(len(headers))])

        # Write to temp file
        temp_file = file_path + ".tmp"
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        for r in rows_to_keep:
            new_sheet.append(r)
        
        new_wb.save(temp_file)
        new_wb.close()
        wb.close()

        # Replace original
        os.remove(file_path)
        os.rename(temp_file, file_path)
        print(f"Successfully cleaned: {file_path} (Removed {sheet.max_row - len(rows_to_keep)} duplicates)")
        
    except Exception as e:
        print(f"Failed to clean {file_path}: {e}")

files = [
    r'd:\projects\Github\Mentoring-Management-System\mock_data\SubjectRegistration_Mock.xlsx',
    r'd:\projects\Github\Mentoring-Management-System\mock_data\marks_upload.xlsx',
    r'd:\projects\Github\Mentoring-Management-System\mock_data\Attendance_Mock_Data.xlsx'
]

for f in files:
    robust_clean(f)
