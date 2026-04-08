import openpyxl
import os

files = [
    r'd:\projects\Github\Mentoring-Management-System\mock_data\marks_upload.xlsx'
]

for f in files:
    if not os.path.exists(f): continue
    wb = openpyxl.load_workbook(f)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    sem_idx = -1
    for i, h in enumerate(headers):
        if h and str(h).lower().strip() in ['semester', 'sem', 'semster']:
            sem_idx = i + 1
            break
    if sem_idx != -1:
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=sem_idx).value = "3rd"
        wb.save(f)
        print(f"Updated {f} to '3rd'.")
    else:
        print(f"Semester column not found in {f}")
