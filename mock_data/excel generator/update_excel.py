import openpyxl
import os

def update_semester(file_path, sheet_name=None):
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    print(f"Updating {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active if sheet_name is None else wb[sheet_name]

    # Find the semester column index
    headers = [cell.value for cell in sheet[1]]
    sem_idx = -1
    for i, h in enumerate(headers):
        if h and str(h).lower().strip() in ['semester', 'sem', 'semster']:
            sem_idx = i + 1
            break

    if sem_idx == -1:
        print(f"Semester column not found in {file_path}. Headers: {headers}")
        return

    # Update all rows
    count = 0
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=sem_idx).value = "3rd"
        count += 1

    wb.save(file_path)
    print(f"Updated {count} rows in {file_path} to '3rd'.")

files = [
    r'd:\projects\Github\Mentoring-Management-System\mock_data\Students_Mock_Data.xlsx',
    r'd:\projects\Github\Mentoring-Management-System\mock_data\SubjectRegistration_Mock.xlsx'
]

for f in files:
    update_semester(f)
