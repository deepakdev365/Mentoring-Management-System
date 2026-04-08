import openpyxl
import random

def generate_backlog_data():
    # 1. Get student data from Students_Mock_Data.xlsx
    students_wb = openpyxl.load_workbook(r'Students_Mock_Data.xlsx')
    students_sheet = students_wb.active
    headers = [cell.value for cell in students_sheet[1]]
    
    reg_idx = -1
    name_idx = -1
    email_idx = -1
    branch_idx = -1
    for i, h in enumerate(headers):
        h_low = str(h).lower().strip()
        if h_low in ['registrationnumber', 'registrationno', 'rollno']:
            reg_idx = i + 1
        elif h_low in ['fullname', 'studentname', 'name']:
            name_idx = i + 1
        elif h_low in ['email', 'mail']:
            email_idx = i + 1
        elif h_low in ['branch', 'department', 'dept']:
            branch_idx = i + 1
    
    if reg_idx == -1: 
        print("Required columns not found in students sheet.")
        return

    student_records = []
    for row in range(2, students_sheet.max_row + 1):
        reg = students_sheet.cell(row=row, column=reg_idx).value
        name = students_sheet.cell(row=row, column=name_idx).value if name_idx != -1 else "Student"
        email = students_sheet.cell(row=row, column=email_idx).value if email_idx != -1 else "email@example.com"
        branch = students_sheet.cell(row=row, column=branch_idx).value if branch_idx != -1 else "CSE"
        if reg: student_records.append({'reg': str(reg), 'name': str(name), 'email': str(email), 'branch': str(branch)})

    # 2. Get subject data from SubjectRegistration_Mock.xlsx
    sub_wb = openpyxl.load_workbook(r'SubjectRegistration_Mock.xlsx')
    sub_sheet = sub_wb.active
    sub_headers = [cell.value for cell in sub_sheet[1]]
    
    code_idx = -1
    sub_name_idx = -1
    for i, h in enumerate(sub_headers):
        h_low = str(h).lower().strip()
        if h_low in ['subjectcode', 'code']:
            code_idx = i + 1
        elif h_low in ['subjectname', 'subject']:
            sub_name_idx = i + 1
            
    subjects = []
    if code_idx != -1 and sub_name_idx != -1:
        for row in range(2, min(30, sub_sheet.max_row + 1)):
            c = sub_sheet.cell(row=row, column=code_idx).value
            n = sub_sheet.cell(row=row, column=sub_name_idx).value
            if c and n: subjects.append({'code': str(c), 'name': str(n)})
    
    if not subjects:
        subjects = [{'code': 'CS101', 'name': 'Data Structures'}, {'code': 'MA201', 'name': 'Mathematics II'}]

    # 3. Generate Backlog_Mock_Data.xlsx
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Backlogs"
    
    # Headers matching BacklogServiceImpl.java findMissingRequiredFields
    headers = ["fullname", "email", "rollNo", "subjectCode", "subjectName", "semester", "branch", "status"]
    sheet.append(headers)

    # Pick ~20% of students to have backlogs
    backlog_students = random.sample(student_records, int(len(student_records) * 0.2))

    for record in backlog_students:
        # 1 backlog per student
        sub = random.choice(subjects)
        sheet.append([
            record['name'],
            record['email'],
            record['reg'],
            sub['code'],
            sub['name'],
            "2nd", # Usually backlogs are from previous semesters
            record['branch'],
            "PENDING"
        ])

    wb.save('Backlog_Mock_Data.xlsx')
    print(f"Generated Backlog_Mock_Data.xlsx with {sheet.max_row - 1} records.")

if __name__ == "__main__":
    generate_backlog_data()
