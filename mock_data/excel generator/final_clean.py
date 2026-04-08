import openpyxl
import os

def normalize(val):
    return str(val).lower().replace(" ", "").replace("_", "").replace("-", "").strip()

def clean_excel(file_path):
    if not os.path.exists(file_path): return
    print(f"Cleaning {file_path}...")
    
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    
    reg_idx = -1
    code_idx = -1
    sem_idx = -1
    
    for i, h in enumerate(headers):
        h_norm = normalize(h)
        if h_norm in ['registrationnumber', 'rollno', 'registrationno', 'regno']: reg_idx = i
        elif h_norm in ['subjectcode', 'code']: code_idx = i
        elif h_norm in ['semester', 'sem']: sem_idx = i
            
    if reg_idx == -1 or code_idx == -1 or sem_idx == -1:
        print(f"Skipping {file_path}: missing columns. Headers were: {headers}")
        return

    seen = set()
    rows_to_keep = [headers]
    dup_count = 0
    
    for row_num in range(2, sheet.max_row + 1):
        reg = str(sheet.cell(row=row_num, column=reg_idx+1).value).strip()
        code = str(sheet.cell(row=row_num, column=code_idx+1).value).strip()
        sem = str(sheet.cell(row=row_num, column=sem_idx+1).value).strip()
        
        key = f"{reg}|{code}|{sem}"
        if key not in seen:
            seen.add(key)
            rows_to_keep.append([sheet.cell(row=row_num, column=c+1).value for c in range(len(headers))])
        else:
            dup_count += 1

    if dup_count > 0:
        # Create new workbook to avoid permission issues during overwrite if possible, 
        # but here we just overwrite the original sheet
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        for r in rows_to_keep:
            new_sheet.append(r)
        
        # Save to a temporary file first
        temp_path = file_path + ".tmp"
        new_wb.save(temp_path)
        new_wb.close()
        wb.close()
        
        os.remove(file_path)
        os.rename(temp_path, file_path)
        print(f"Removed {dup_count} duplicates from {file_path}.")
    else:
        print(f"No duplicates in {file_path}.")

files = [
    r'd:\projects\Github\Mentoring-Management-System\mock_data\SubjectRegistration_Mock.xlsx',
    r'd:\projects\Github\Mentoring-Management-System\mock_data\marks_upload.xlsx',
    r'd:\projects\Github\Mentoring-Management-System\mock_data\Attendance_Mock_Data.xlsx',
    r'd:\projects\Github\Mentoring-Management-System\mock_data\Payment_Mock_Data.xlsx'
]

for f in files:
    clean_excel(f)
